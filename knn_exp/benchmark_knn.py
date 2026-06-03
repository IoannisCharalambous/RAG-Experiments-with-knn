import numpy as np
import faiss
from annoy import AnnoyIndex
import os
import csv
import pickle
from sentence_transformers import SentenceTransformer
import time
import sys
import math
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from scipy.spatial.distance import pdist

def compute_internal_similarity(row_indices, embeddings):
    valid_indices = [i for i in row_indices if i != -1]
    
    if len(valid_indices) < 2:
        return 0.0
    
    vecs = embeddings[valid_indices]
    distances = pdist(vecs, metric='euclidean')
    maxdistance = max(distances)
    
    return np.sum(distances) / (maxdistance * (pow(len(valid_indices), 2) / 2))

class DataManager:
    def __init__(self, base_dir):
        self.dataset_dir = os.path.join(base_dir, 'datasets', 'dbpedia', 'embeddings')
        self.csv_path = os.path.join(self.dataset_dir, 'dbpedia_popular_os_regions.csv')
        self.embeddings_path = os.path.join(self.dataset_dir, 'embeddings.npy')
        self.ids_path = os.path.join(self.dataset_dir, 'keyword_ids.npy')
        self.keywords_path = os.path.join(self.dataset_dir, 'keyword.txt')
        self.annoy_index_file = os.path.join(self.dataset_dir, 'annoy_index.ann')
        self.annoy_mapping_file = os.path.join(self.dataset_dir, 'annoy_id_map.pkl')
        
    def load_queries(self):
        """Return the KEYWORDS list as query strings.
        Each keyword is encoded as a query vector for KNN search."""
        
        KEYWORDS = [
            
            "Socrates",
            "Aristotle",
            "Pericles",
            "Acropolis",
            "Parthenon",
            
            # "Apple",
            # "Apple TV",
            # "NCAA",
            # "FIFA",

        ]
        
        return KEYWORDS
        
    def load_id_to_text(self):
        id_to_text = {}
        if os.path.exists(self.keywords_path):
            with open(self.keywords_path, 'r', encoding='utf-8') as f:
                for line in f:
                    parts = line.strip().split(maxsplit=1)
                    if len(parts) == 2:
                        id_to_text[int(parts[0])] = parts[1]
        return id_to_text
        
    def load_embeddings_and_ids(self):
        embeddings = np.load(self.embeddings_path, mmap_mode='r')
        node_ids = np.load(self.ids_path)
        return embeddings, node_ids

class Evaluator:
    def __init__(self, results_dir, id_to_text, k_values):
        self.results_dir = results_dir
        self.id_to_text = id_to_text
        self.k_values = k_values
        self.metrics = {} # {algo: {k: {query: {'dist': x, 'time': y, 'internal_sim': z}}}}
        self.raw_results = {} # {algo: {k: {query: [(matched_id, distance), ...], 'time': ms, 'internal_sim': s}}}
        self.ivf_metadata = {}
        self.datasets = {} # Map query -> dataset name
        
    def record_results(self, algo_name, k, queries, results, dataset_name="DBpedia"):
        if algo_name not in self.metrics:
            self.metrics[algo_name] = {}
        if k not in self.metrics[algo_name]:
            self.metrics[algo_name][k] = {}
        
        if algo_name not in self.raw_results:
            self.raw_results[algo_name] = {}
        if k not in self.raw_results[algo_name]:
            self.raw_results[algo_name][k] = {}
        
        for q_idx, query in enumerate(queries):
            matched_ids, distances, q_time_ms, internal_sim = results[q_idx]
            
            sum_dist = 0
            for rank, (matched_id, dist) in enumerate(zip(matched_ids, distances)):
                if matched_id != -1:
                    sum_dist += dist
                    
            self.metrics[algo_name][k][query] = {
                'distance_sum': sum_dist, 
                'time': q_time_ms,
                'internal_sim': internal_sim
            }
            self.datasets[query] = dataset_name
            
            # Store raw per-rank results for CSV export
            self.raw_results[algo_name][k][query] = {
                'matches': list(zip(matched_ids, distances)),
                'time': q_time_ms,
                'internal_sim': internal_sim
            }

    def generate_summary(self, queries):
        self._generate_csvs(queries)
        self._generate_markdown(queries)
        self._generate_excel(queries)
    
    def _generate_csvs(self, queries):
        """Write per-algorithm, per-k CSV result files.
        Each CSV has columns: Query, Rank, Matched Keyword, Matched ID,
        Distance, Search Time (ms), Internal Similarity
        """
        for algo in self.raw_results:
            for k in self.k_values:
                if k not in self.raw_results[algo]:
                    continue
                
                csv_path = os.path.join(
                    self.results_dir, f'{algo}_k{k}_results.csv'
                )
                
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow([
                        'Query', 'Rank', 'Matched Keyword', 'Matched ID',
                        'Distance', 'Search Time (ms)', 'Internal Similarity'
                    ])
                    
                    for q_idx, query in enumerate(queries):
                        if query not in self.raw_results[algo][k]:
                            continue
                        
                        # Add blank separator row between keyword groups
                        if q_idx > 0:
                            writer.writerow([])
                        
                        entry = self.raw_results[algo][k][query]
                        q_time = entry['time']
                        int_sim = entry['internal_sim']
                        
                        for rank, (matched_id, dist) in enumerate(entry['matches'], 1):
                            keyword = self.id_to_text.get(
                                int(matched_id), str(matched_id)
                            ) if matched_id != -1 else 'N/A'
                            
                            writer.writerow([
                                query, rank, keyword,
                                int(matched_id) if matched_id != -1 else -1,
                                f'{dist:.6f}', f'{q_time:.3f}',
                                f'{int_sim:.6f}'
                            ])
                
                print(f"  Written {csv_path}")
        
    def _generate_excel(self, queries):
        base_algo = 'faiss_l2'
        if base_algo not in self.metrics:
            return
            
        algos = list(self.metrics.keys())
        diff_rows = []
        
        prev_dataset = None
        for k in self.k_values:
            for query in queries:
                if query not in self.metrics[base_algo][k]:
                    continue
                    
                dataset = self.datasets.get(query, "Unknown")
                
                marker = False
                if prev_dataset is not None and dataset != prev_dataset:
                    marker = True
                prev_dataset = dataset
                
                base_time = self.metrics[base_algo][k][query]['time']
                base_dist = self.metrics[base_algo][k][query]['distance_sum']
                
                row = {
                    'Dataset': dataset,
                    'Marker_NextDataset': marker,
                    'K': k,
                    'Query': query
                }
                
                for algo in algos:
                    if query in self.metrics[algo][k]:
                        t = self.metrics[algo][k][query]['time']
                        d = self.metrics[algo][k][query]['distance_sum']
                        s = self.metrics[algo][k][query]['internal_sim']
                        
                        row[f'{algo} time'] = t
                        row[f'{algo} distance (from query)'] = d
                        row[f'{algo} similarity'] = s
                        
                        if algo != base_algo:
                            speedup = base_time / t if t > 0 else 1
                            pct_diff = ((d - base_dist) / base_dist * 100) if base_dist > 0 else 0
                            row[f'{algo} speedup'] = speedup
                            row[f'{algo} diff%'] = pct_diff
                diff_rows.append(row)
                
        df_diffs = pd.DataFrame(diff_rows)
        
        # Strict constraint: Drop "shape" column if it somehow exists
        if 'shape' in df_diffs.columns:
            df_diffs.drop(columns=['shape'], inplace=True)
            
        summary_rows = []
        for k in self.k_values:
            df_k = df_diffs[df_diffs['K'] == k]
            if df_k.empty:
                continue
                
            agg_funcs = {
                'MEAN': df_k.mean(numeric_only=True), 
                'MAX': df_k.max(numeric_only=True), 
                'MIN': df_k.min(numeric_only=True), 
                'STD': df_k.std(numeric_only=True)
            }
            
            for t_name, agg_series in agg_funcs.items():
                s_row = {'K': k, 'Type': t_name}
                for col in df_diffs.columns:
                    if col not in ['Dataset', 'Marker_NextDataset', 'K', 'Query', 'Type', 'shape']:
                        s_row[col] = agg_series.get(col, np.nan)
                summary_rows.append(s_row)
                
        df_summary = pd.DataFrame(summary_rows)
        if 'shape' in df_summary.columns:
            df_summary.drop(columns=['shape'], inplace=True)
            
        out_path = os.path.join(self.results_dir, 'benchmark_results.xlsx')
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            diff_sheet = writer.book.create_sheet('Diffs')
            diff_cols = [c for c in df_diffs.columns if c != 'Marker_NextDataset']
            diff_sheet.append(diff_cols)
            
            for _, row in df_diffs.iterrows():
                if row['Marker_NextDataset']:
                    diff_sheet.append(["NEXT QUERY"] + [""]*(len(diff_cols)-1))
                diff_sheet.append([row[c] for c in diff_cols])
                
        wb = openpyxl.load_workbook(out_path)
        self._apply_pro_styling(wb['Summary'], algos, is_summary=True)
        self._apply_pro_styling(wb['Diffs'], algos, is_summary=False)
        wb.save(out_path)

    def _apply_pro_styling(self, ws, algos, is_summary):
        header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin', color='BFBFBF'), 
                             right=Side(style='thin', color='BFBFBF'), 
                             top=Side(style='thin', color='BFBFBF'), 
                             bottom=Side(style='thin', color='BFBFBF'))
                             
        algo_colors = {'faiss_l2': 'E2EFDA', 'faiss_ivf': 'DDEBF7', 'annoy': 'FFF2CC'}
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        ws.freeze_panes = "A2"
        
        col_algo_map = {}
        header_row = [c.value for c in ws[1]]
        for idx, col_name in enumerate(header_row, 1):
            if not col_name: continue
            for algo in algos:
                if algo in str(col_name).lower() or algo.replace('_', ' ') in str(col_name).lower():
                    col_algo_map[idx] = algo_colors.get(algo, 'FFFFFF')
                    break

        def apply_sig_fig_format(c, sig=3, min_decimals=6):
            val = c.value
            if isinstance(val, (int, float)) and not pd.isna(val):
                if val == 0:
                    c.number_format = "0." + "0" * min_decimals
                else:
                    try:
                        mag = int(math.floor(math.log10(abs(val))))
                        decimals = max(min_decimals, sig - mag - 1)
                        fmt = "0." + "0" * decimals
                        c.number_format = fmt
                    except:
                        pass

        for row in ws.iter_rows(min_row=2):
            is_next_query = (row[0].value == "NEXT QUERY")
            if is_next_query:
                for cell in row:
                    cell.font = Font(color="FF0000", bold=True)
                    cell.alignment = Alignment(horizontal="left")
                continue
                
            row_type = None
            if is_summary:
                row_type = row[1].value 
                
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
                    
                if cell.column in col_algo_map and cell.value is not None:
                    cell.fill = PatternFill(start_color=col_algo_map[cell.column], end_color=col_algo_map[cell.column], fill_type="solid")
                    
                if is_summary and row_type and cell.value is not None:
                    if row_type == 'MEAN':
                        cell.font = Font(color="1F4E78", bold=True)
                    elif row_type in ['MAX', 'MIN']:
                        cell.font = Font(color="833C0C", italic=True)
                    elif row_type == 'STD':
                        cell.font = Font(color="9C5700", italic=True)

                col_name = str(header_row[cell.column - 1]) if cell.column <= len(header_row) else ""
                if 'time' in col_name.lower() or 'speedup' in col_name.lower() or 'latency' in col_name.lower():
                    apply_sig_fig_format(cell, 3)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    def _generate_markdown(self, queries):
        md_path = os.path.join(self.results_dir, 'comparison_summary.md')
        
        base_algo = 'faiss_l2'
        if base_algo not in self.metrics:
            return
            
        try:
            import matplotlib.pyplot as plt
            
            labels = []
            data = []
            for algo in self.metrics:
                for k in self.k_values:
                    times = [self.metrics[algo][k][q]['time'] for q in queries if q in self.metrics[algo][k]]
                    if times:
                        labels.append(f"{algo}\nk={k}")
                        data.append(times)
                        
            plt.figure(figsize=(14, 7))
            plt.boxplot(data, labels=labels)
            plt.yscale('log')
            plt.ylabel('Search Time (ms) - Log Scale')
            plt.title('Search Latency Distribution by Algorithm and K')
            plt.xticks(rotation=45)
            plt.tight_layout()
            plot_path = os.path.join(self.results_dir, 'latency_distribution.png')
            plt.savefig(plot_path)
            plt.close()
        except ImportError:
            print("matplotlib not installed. Skipping latency plot generation.")

        with open(md_path, 'w', encoding='utf-8') as f:
            f.write("# KNN Algorithms Comparison Summary\n\n")
            
            if self.ivf_metadata:
                f.write("## FAISS IVF Configuration\n")
                f.write(f"- **nlist (Clusters Trained)**: {self.ivf_metadata.get('nlist', 'N/A')}\n")
                f.write(f"- **nprobe (Clusters Searched)**: {self.ivf_metadata.get('nprobe', 'N/A')}\n\n")
                
            plot_path = os.path.join(self.results_dir, 'latency_distribution.png')
            if os.path.exists(plot_path):
                f.write("## Latency Distribution\n")
                f.write("![Latency Distribution](latency_distribution.png)\n\n")
            
            f.write("## Results per K\n\n")
            algos = list(self.metrics.keys())
            
            for k in self.k_values:
                f.write(f"### K = {k}\n\n")
                
                f.write("| Query | ")
                for algo in algos:
                    f.write(f"{algo} L2 Sum | {algo} Time (ms) | {algo} Avg Int. Sim | ")
                    if algo != base_algo:
                        f.write(f"% L2 Diff | Speedup | ")
                f.write("\n| :--- | ")
                for algo in algos:
                    f.write(":---: | :---: | :---: | ")
                    if algo != base_algo:
                        f.write(":---: | :---: | ")
                f.write("\n")
                
                global_stats = {algo: {'pct_diff': 0, 'time': 0, 'int_sim': 0, 'count': 0} for algo in algos}
                
                for query in queries:
                    f.write(f"| {query} | ")
                    
                    if query not in self.metrics[base_algo][k]:
                        f.write("\n")
                        continue
                        
                    base_dist = self.metrics[base_algo][k][query]['distance_sum']
                    base_time = self.metrics[base_algo][k][query]['time']
                    
                    for algo in algos:
                        dist = self.metrics[algo][k][query]['distance_sum']
                        time_ms = self.metrics[algo][k][query]['time']
                        int_sim = self.metrics[algo][k][query]['internal_sim']
                        
                        f.write(f"{dist:.4f} | {time_ms:.3f} | {int_sim:.4f} | ")
                        
                        global_stats[algo]['time'] += time_ms
                        global_stats[algo]['int_sim'] += int_sim
                        global_stats[algo]['count'] += 1
                        
                        if algo != base_algo:
                            pct_diff = ((dist - base_dist) / base_dist * 100) if base_dist > 0 else 0
                            speedup = base_time / time_ms if time_ms > 0 else 1
                            f.write(f"{pct_diff:.2f}% | {speedup:.2f}x | ")
                            global_stats[algo]['pct_diff'] += pct_diff
                    f.write("\n")
                    
                f.write("\n**Summary Statistics (k={})**\n\n".format(k))
                base_avg_time = global_stats[base_algo]['time'] / max(1, global_stats[base_algo]['count'])
                base_avg_sim = global_stats[base_algo]['int_sim'] / max(1, global_stats[base_algo]['count'])
                f.write(f"- **{base_algo}** - Avg Time: {base_avg_time:.3f} ms, Avg Internal Sim: {base_avg_sim:.4f}\n")
                
                for algo in algos:
                    if algo == base_algo: continue
                    avg_pct_diff = global_stats[algo]['pct_diff'] / max(1, global_stats[algo]['count'])
                    avg_time = global_stats[algo]['time'] / max(1, global_stats[algo]['count'])
                    avg_sim = global_stats[algo]['int_sim'] / max(1, global_stats[algo]['count'])
                    speedup = base_avg_time / avg_time if avg_time > 0 else 1
                    f.write(f"- **{algo}** - Avg Time: {avg_time:.3f} ms (Speedup: {speedup:.2f}x), Avg Internal Sim: {avg_sim:.4f}, Avg Accuracy Loss: {avg_pct_diff:.2f}%\n")
                f.write("\n---\n\n")

def run_faiss_ivf(embeddings, query_vectors, node_ids, k_values, evaluator):
    num_vectors, dimension = embeddings.shape
    nlist = int(4 * np.sqrt(num_vectors))
    nprobe = 10
    
    evaluator.ivf_metadata = {'nlist': nlist, 'nprobe': nprobe}
    
    quantizer = faiss.IndexFlatL2(dimension)
    index = faiss.IndexIVFFlat(quantizer, dimension, nlist, faiss.METRIC_L2)
    
    print(f"  Training IVF with nlist={nlist}...")
    train_size = min(100000, num_vectors)
    train_indices = np.random.choice(num_vectors, train_size, replace=False)
    train_data = np.array([embeddings[i] for i in train_indices]) 
    index.train(train_data)
    
    print("  Adding vectors to IVF index...")
    chunk_size = 500000
    for i in range(0, num_vectors, chunk_size):
        end = min(i + chunk_size, num_vectors)
        index.add(np.array(embeddings[i:end]))
        
    index.nprobe = nprobe
    
    all_results = {}
    for k in k_values:
        print(f"  Evaluating K={k}...")
        results = []
        for q_vec in query_vectors:
            t0 = time.perf_counter()
            distances, indices = index.search(np.array([q_vec]), k)
            q_time_ms = (time.perf_counter() - t0) * 1000
            euclidean_distances = [math.sqrt(max(0, d)) for d in distances[0]]
            matched_ids = [node_ids[i] if i != -1 else -1 for i in indices[0]]
            int_sim = compute_internal_similarity(indices[0], embeddings)
            results.append((matched_ids, euclidean_distances, q_time_ms, int_sim))
        all_results[k] = results
    return all_results

def run_annoy(embeddings, query_vectors, node_ids, annoy_index_path, annoy_map_path, k_values):
    dimension = embeddings.shape[1]
    num_vectors = embeddings.shape[0]
    index = AnnoyIndex(dimension, 'euclidean')
    
    id_mapping = {}
    if os.path.exists(annoy_index_path) and os.path.exists(annoy_map_path):
        print("  Loading existing Annoy index...")
        index.load(annoy_index_path)
        with open(annoy_map_path, 'rb') as f:
            id_mapping = pickle.load(f)
    else:
        print("  Building Annoy index (this may take a while)...")
        for i in range(num_vectors):
            index.add_item(i, embeddings[i])
            id_mapping[i] = node_ids[i]
            if i > 0 and i % 500000 == 0:
                print(f"    Added {i:,} vectors...")
        index.build(50, n_jobs=-1)
        index.save(annoy_index_path)
        with open(annoy_map_path, 'wb') as f:
            pickle.dump(id_mapping, f)
            
    all_results = {}
    for k in k_values:
        print(f"  Evaluating K={k}...")
        results = []
        for q_vec in query_vectors:
            t0 = time.perf_counter()
            indices, distances = index.get_nns_by_vector(q_vec, k, search_k=50000, include_distances=True)
            q_time_ms = (time.perf_counter() - t0) * 1000
            matched_ids = [id_mapping[i] for i in indices]
            int_sim = compute_internal_similarity(indices, embeddings)
            results.append((matched_ids, distances, q_time_ms, int_sim))
        all_results[k] = results
    return all_results

def main():
    if sys.platform.startswith('win'):
        sys.stdout.reconfigure(encoding='utf-8')
        
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(script_dir)
    
    dm = DataManager(base_dir)
    print("Loading queries...")
    queries = dm.load_queries()
    print(f"Generated {len(queries)} queries.")
    
    print("Loading id to text mapping...")
    id_to_text = dm.load_id_to_text()
    
    print("Loading embeddings and ids...")
    embeddings, node_ids = dm.load_embeddings_and_ids()
    print(f"Embeddings shape: {embeddings.shape}")
    
    print("Loading sentence transformer model...")
    model = SentenceTransformer('all-MiniLM-L6-v2', device='cuda' if faiss.get_num_gpus() > 0 else 'cpu')
    query_vectors = model.encode(queries, convert_to_numpy=True)
    
    k_values = [5, 20, 50, 100]
    evaluator = Evaluator(script_dir, id_to_text, k_values)
    
    print("\nRunning FAISS L2 Exact...")
    dimension = embeddings.shape[1]
    num_vectors = embeddings.shape[0]
    l2_index = faiss.IndexFlatL2(dimension)
    chunk_size = 500000
    for i in range(0, num_vectors, chunk_size):
        end = min(i + chunk_size, num_vectors)
        l2_index.add(np.array(embeddings[i:end]))
        
    for k in k_values:
        print(f"  Evaluating K={k}...")
        l2_results = []
        for q_vec in query_vectors:
            t0 = time.perf_counter()
            distances, indices = l2_index.search(np.array([q_vec]), k)
            q_time_ms = (time.perf_counter() - t0) * 1000
            euclidean_distances = [math.sqrt(max(0, d)) for d in distances[0]]
            matched_ids = [node_ids[i] for i in indices[0]]
            int_sim = compute_internal_similarity(indices[0], embeddings)
            l2_results.append((matched_ids, euclidean_distances, q_time_ms, int_sim))
        evaluator.record_results('faiss_l2', k, queries, l2_results)
    
    print("\nRunning FAISS IVF...")
    ivf_all_results = run_faiss_ivf(embeddings, query_vectors, node_ids, k_values, evaluator)
    for k in k_values:
        evaluator.record_results('faiss_ivf', k, queries, ivf_all_results[k])
    
    print("\nRunning Annoy...")
    annoy_all_results = run_annoy(embeddings, query_vectors, node_ids, dm.annoy_index_file, dm.annoy_mapping_file, k_values)
    for k in k_values:
        evaluator.record_results('annoy', k, queries, annoy_all_results[k])
    
    print("\nGenerating summary...")
    evaluator.generate_summary(queries)
    print("Done! Check knn_exp directory for results.")

if __name__ == "__main__":
    main()
